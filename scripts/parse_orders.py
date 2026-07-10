#!/usr/bin/env python3
"""
~/Dropbox/중요 폴더/발주서/ 안의 xlsx 발주서를 파싱해서
Brightbeed data/orders.json 생성.

파일명 규칙: "MM.DD 발주서.xlsx" 또는 "MM.DD 발주서 (설명).xlsx"
  - 년도는 파일명에 없음 → mtime의 년도를 사용 (월이 안맞으면 -1년 조정)
  - 괄호 안 문구는 note로 추출 (예: 케이블 파우치, 식스에이 must, 6a, ct 등)

헤더는 대부분 R4:
  신형: 상품링크|상품사진|장바구니단가|제품명 및 옵션명|1688 옵션명|메모(소요일)|수량|1688단가￥|중국물류비|합계￥|적용환율|상품가|구매대행수수료|수수료 세금계산서|총입금하실금액￦
  구형: 상품링크|상품사진|한글옵션명|1688옵션명|메모|수량|1688단가￥|...
"""

import json
import os
import re
import sys
import unicodedata
from datetime import datetime, date
from pathlib import Path

import openpyxl

FOLDER = Path(os.path.expanduser("~/Dropbox/중요 폴더/발주서"))
OUT = Path(__file__).resolve().parent.parent / "data" / "orders.json"


def nfc(s):
    return unicodedata.normalize("NFC", s) if isinstance(s, str) else s


# 파일명 앞쪽: MM.DD (예: "01.07 발주서")
FNAME_MMDD = re.compile(r"^\s*(\d{1,2})\.(\d{1,2})\b")
# YYMMDD 6자리 (예: 6A 240624, Sb260609)
FNAME_YYMMDD = re.compile(r"(?<!\d)(\d{2})(\d{2})(\d{2})(?!\d)")
# YY.MM.DD (예: 브라이트비드 발주서_제이 25.04.08)
FNAME_YYDMDMD = re.compile(r"(?<!\d)(\d{2})\.(\d{1,2})\.(\d{1,2})(?!\d)")
# 괄호 안 텍스트
PAREN_NOTE = re.compile(r"\(([^)]+)\)")
IS_EXTRA = re.compile(r"추가")


def parse_mmdd(name):
    m = FNAME_MMDD.match(name)
    if not m:
        return None
    mm, dd = int(m.group(1)), int(m.group(2))
    if 1 <= mm <= 12 and 1 <= dd <= 31:
        return mm, dd
    return None


def parse_yymmdd_anywhere(name):
    """YY.MM.DD 우선, 아니면 YYMMDD 6자리 매칭."""
    m = FNAME_YYDMDMD.search(name)
    if m:
        yy, mm, dd = int(m.group(1)), int(m.group(2)), int(m.group(3))
        if 20 <= yy <= 30 and 1 <= mm <= 12 and 1 <= dd <= 31:
            return 2000 + yy, mm, dd
    m = FNAME_YYMMDD.search(name)
    if m:
        yy, mm, dd = int(m.group(1)), int(m.group(2)), int(m.group(3))
        if 20 <= yy <= 30 and 1 <= mm <= 12 and 1 <= dd <= 31:
            return 2000 + yy, mm, dd
    return None


def resolve_date(name: str, mtime: float) -> str:
    """파일명에서 날짜 최대한 뽑고, 안되면 mtime."""
    # 1) YY.MM.DD 또는 YYMMDD (파일명 어디서든)
    ymd = parse_yymmdd_anywhere(name)
    if ymd:
        y, m, d = ymd
        try:
            return date(y, m, d).isoformat()
        except ValueError:
            pass
    # 2) MM.DD 접두어 (년도는 mtime에서)
    mmdd = parse_mmdd(name)
    if mmdd:
        mm, dd = mmdd
        mtime_dt = datetime.fromtimestamp(mtime)
        year = mtime_dt.year
        # 파일명 월-mtime 월 6개월 이상 → 년도 조정
        if mm - mtime_dt.month > 6:
            year -= 1
        elif mtime_dt.month - mm > 6:
            year += 1
        try:
            return date(year, mm, dd).isoformat()
        except ValueError:
            pass
    # 3) mtime 폴백
    return datetime.fromtimestamp(mtime).date().isoformat()


def extract_note(name: str) -> str:
    """파일명 괄호 안 첫 항목만 반환 (Q번호·전승훈·전수빈 같은 반복 값 제거)."""
    for m in PAREN_NOTE.finditer(name):
        val = m.group(1).strip()
        # 반복되는 담당자·업체 코드 스킵
        if re.search(r"Q\d{4}", val) or val in ("1", "2", "3", "4"):
            continue
        if val:
            return val
    return ""


def extract_group(name: str) -> str:
    """파일명 접두어로 발주처/카테고리 그룹 판별."""
    n = name.strip()
    # 발주서 아닌 것: 출고/남은수량/재고/개별발송/양식/참고
    if any(k in n for k in ("출고", "남은수량", "개별발송", "양식", "브랜드별", "SA컴퍼니")) or n.startswith("케이블 주문 수량"):
        return "재고·양식"
    if n.startswith("6A") or "Q2215-6A" in n:
        return "6A(식스에이)"
    if n.startswith("Sb") or n.startswith("SB") or "Q2365" in n:
        return "샤인비드"
    if "브라이트비드 발주서_제이" in n:
        return "제이"
    if n.startswith("케이블"):
        return "케이블"
    if re.match(r"^\d{1,2}\.\d{1,2}", n):
        return "일반"
    return "기타"


def is_extra(name: str) -> bool:
    return bool(IS_EXTRA.search(name))


def to_num(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).replace(",", "").replace("￦", "").replace("￥", "").replace("원", "").strip()
    try:
        return float(s)
    except Exception:
        return None


SUMMARY_LABELS = ("합계", "총수량", "총 수량", "총입금", "총 입금", "상품가", "구매대행수수료",
                  "세금계산서", "적용환율")


def is_summary_row(row_s):
    """행의 어떤 셀이든 요약 라벨 텍스트를 포함하면 True."""
    joined = " ".join(row_s[:15])
    for label in SUMMARY_LABELS:
        if label in joined:
            return True
    return False


HEADER_ALIASES = {
    "link": ["상품링크"],
    "img": ["상품이미지", "상품사진"],
    "basket_price": ["장바구니단가"],
    "kor": ["제품명및옵션명", "한글옵션명", "제품명옵션명"],
    "cn": ["1688옵션명", "1668옵션명"],
    "memo": ["메모", "메모소요일"],
    "qty": ["수량"],
    "price_cny": ["1688단가", "1688단가￥", "단가"],
    "ship_cny": ["중국물류비"],
}


def find_header_row(ws):
    """1~10 행 안에서 '상품링크' + '수량' 을 포함한 행을 찾아 (row_idx, col_map) 반환."""
    for r_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=10, max_col=20, values_only=True), start=1):
        row_s = [nfc(str(v)).replace("\n", "").replace(" ", "").strip() if v is not None else "" for v in row]
        if "상품링크" in row_s and "수량" in row_s:
            col_map = {}
            for i, s in enumerate(row_s):
                if not s:
                    continue
                for canon, aliases in HEADER_ALIASES.items():
                    if canon in col_map:
                        continue
                    for a in aliases:
                        if s == a or s.startswith(a):
                            col_map[canon] = i
                            break
            return r_idx, col_map
    return None, {}


def parse_xls_via_xlrd(path: Path):
    """구버전 .xls는 xlrd 로 파싱."""
    try:
        import xlrd
    except ImportError:
        return None
    try:
        book = xlrd.open_workbook(str(path))
    except Exception:
        return None
    ws = book.sheet_by_index(0)
    # 헤더 찾기 (1~10행)
    col_map = {}
    header_row = None
    for r in range(min(10, ws.nrows)):
        row_s = []
        for c in range(min(20, ws.ncols)):
            v = ws.cell_value(r, c)
            row_s.append(nfc(str(v)).replace("\n", "").replace(" ", "").strip() if v not in (None, "") else "")
        if "상품링크" in row_s and "수량" in row_s:
            header_row = r
            for i, s in enumerate(row_s):
                if not s:
                    continue
                for canon, aliases in HEADER_ALIASES.items():
                    if canon in col_map:
                        continue
                    for a in aliases:
                        if s == a or s.startswith(a):
                            col_map[canon] = i
                            break
            break
    items = []
    total_cny = 0.0
    total_cny_any = False
    last_link_x = ""
    last_img_x = ""
    if header_row is not None:
        c_link = col_map.get("link")
        c_img = col_map.get("img")
        c_kor = col_map.get("kor")
        c_cn = col_map.get("cn")
        c_memo = col_map.get("memo")
        c_qty = col_map.get("qty")
        c_price = col_map.get("price_cny")
        c_basket = col_map.get("basket_price")
        for r in range(header_row + 1, min(header_row + 501, ws.nrows)):
            row = [ws.cell_value(r, c) for c in range(min(20, ws.ncols))]
            row_s = [nfc(str(v)).strip() if v not in (None, "") else "" for v in row]

            def get(cidx, limit=200):
                if cidx is None or cidx >= len(row_s):
                    return ""
                v = row_s[cidx]
                if v.startswith("=") or v.startswith("_xlfn."):
                    return ""
                return v[:limit]

            qty = to_num(row[c_qty]) if c_qty is not None and c_qty < len(row) else None
            price = to_num(row[c_price]) if c_price is not None and c_price < len(row) else None
            if price is None and c_basket is not None and c_basket < len(row):
                price = to_num(row[c_basket])
            kor = get(c_kor)
            cn = get(c_cn)
            link = get(c_link, 400)
            if is_summary_row(row_s):
                break
            if not (link or kor or cn):
                continue
            img = get(c_img, 400)
            if not img or not img.startswith("http"):
                img = ""
                for j in range(min(10, len(row_s))):
                    v = row_s[j]
                    if v.startswith("http") and "alicdn" in v and (c_link is None or j != c_link):
                        img = v[:400]
                        break
            if not link and last_link_x:
                link = last_link_x
            else:
                last_link_x = link
            if not img and last_img_x:
                img = last_img_x
            else:
                last_img_x = img
            items.append({
                "img": img, "link": link, "kor": kor, "cn": cn,
                "memo": get(c_memo), "qty": qty, "price_cny": price,
            })
            if qty is not None and price is not None:
                total_cny += qty * price
                total_cny_any = True
    return items, (round(total_cny, 2) if total_cny_any else None)


def parse_workbook(path: Path):
    # .xls 는 xlrd 로 (openpyxl 로 못 읽음)
    if path.suffix.lower() == ".xls":
        pass  # 아래서 xlrd 폴백
    try:
        if path.suffix.lower() == ".xls":
            raise ValueError("xls: use xlrd")
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    except Exception as e:
        fname = nfc(path.name)
        stat = path.stat()
        base = {
            "file": fname,
            "size": stat.st_size,
            "mtime": datetime.fromtimestamp(stat.st_mtime).isoformat(timespec="seconds"),
            "date": resolve_date(fname, stat.st_mtime),
            "group": extract_group(fname),
            "note": extract_note(fname),
            "extra": is_extra(fname),
            "items": [],
            "item_count": 0,
            "total_qty": None,
            "total_cny": None,
        }
        xlrd_res = parse_xls_via_xlrd(path)
        if xlrd_res is not None:
            items, total_cny = xlrd_res
            base["items"] = items
            base["item_count"] = len(items)
            qtys = [it["qty"] for it in items if it["qty"] is not None]
            if qtys:
                base["total_qty"] = sum(qtys)
            base["total_cny"] = total_cny
            return base
        return {"error": f"open failed: {e}"}

    fname = nfc(path.name)
    stat = path.stat()
    result = {
        "file": fname,
        "size": stat.st_size,
        "mtime": datetime.fromtimestamp(stat.st_mtime).isoformat(timespec="seconds"),
        "date": resolve_date(fname, stat.st_mtime),
        "group": extract_group(fname),
        "note": extract_note(fname),
        "extra": is_extra(fname),
        "items": [],
        "item_count": 0,
        "total_qty": None,
        "total_cny": None,
    }

    ws = wb[wb.sheetnames[0]]
    header_row, col_map = find_header_row(ws)
    if header_row is None:
        try:
            wb.close()
        except Exception:
            pass
        return result

    c_link = col_map.get("link")
    c_img = col_map.get("img")
    c_kor = col_map.get("kor")
    c_cn = col_map.get("cn")
    c_memo = col_map.get("memo")
    c_qty = col_map.get("qty")
    c_price = col_map.get("price_cny")
    c_basket = col_map.get("basket_price")

    MAX_ROWS = 500
    total_cny = 0.0
    total_cny_any = False
    last_link = ""
    last_img = ""

    for r_idx, row in enumerate(ws.iter_rows(min_row=header_row + 1, max_row=header_row + MAX_ROWS, max_col=20, values_only=True), start=header_row + 1):
        row = list(row)
        row_s = [nfc(str(v)).strip() if v is not None else "" for v in row]

        def get(cidx, limit=200):
            if cidx is None or cidx >= len(row_s):
                return ""
            v = row_s[cidx]
            # =DISPIMG(...) 같은 엑셀 수식은 상품명으로 부적절
            if v.startswith("=") or v.startswith("_xlfn."):
                return ""
            return v[:limit]

        qty = to_num(row[c_qty]) if c_qty is not None and c_qty < len(row) else None
        price = to_num(row[c_price]) if c_price is not None and c_price < len(row) else None
        if price is None and c_basket is not None and c_basket < len(row):
            # 장바구니단가만 있고 1688단가 비어있으면 대체
            price = to_num(row[c_basket])
        kor = get(c_kor)
        cn = get(c_cn)
        link = get(c_link, 400)

        # 요약 라벨 (합계·총수량·총입금·상품가 등)이 어디든 있으면 하단 요약 시작 → 끝
        if is_summary_row(row_s):
            break
        # 상품 정보 (링크/한글명/1688명) 하나라도 있어야 아이템으로 인정
        if not (link or kor or cn):
            continue

        img = get(c_img, 400)
        if not img or not img.startswith("http"):
            # 폴백: 첫 10칸 안에서 상품링크와 별개의 alicdn 이미지 URL 탐색
            img = ""
            for j in range(min(10, len(row_s))):
                v = row_s[j]
                if v.startswith("http") and "alicdn" in v and (c_link is None or j != c_link):
                    img = v[:400]
                    break
        # 색상별 연속 행: link/img 비어있으면 이전 행 값 상속
        if not link and last_link:
            link = last_link
        else:
            last_link = link
        if not img and last_img:
            img = last_img
        else:
            last_img = img
        item = {
            "img": img,
            "link": link,
            "kor": kor,
            "cn": cn,
            "memo": get(c_memo),
            "qty": qty,
            "price_cny": price,
        }
        result["items"].append(item)
        if qty is not None and price is not None:
            total_cny += qty * price
            total_cny_any = True

    result["item_count"] = len(result["items"])
    qtys = [it["qty"] for it in result["items"] if it["qty"] is not None]
    if qtys:
        result["total_qty"] = sum(qtys)
    if total_cny_any:
        result["total_cny"] = round(total_cny, 2)

    try:
        wb.close()
    except Exception:
        pass
    return result


DUP_TAIL = re.compile(r"\s*\(\d+\)\s*$")


def dedupe_key(fname: str) -> str:
    """중복 감지용 정규화 파일명.
    - 확장자 대소문자 통일
    - 끝의 (1)/(2)/(3) 반복 제거
    - 여러 공백 → 단일 공백
    """
    fname = nfc(fname)
    stem, ext = os.path.splitext(fname)
    prev = None
    while prev != stem:
        prev = stem
        stem = DUP_TAIL.sub("", stem).strip()
    stem = re.sub(r"\s+", " ", stem).strip()
    return stem + ext.lower()


_OFFER_RE = re.compile(r"/offer/(\d+)")
_IBANK_RE = re.compile(r"/img/ibank/(O1CN[0-9A-Za-z]+)")


def _norm_link(url):
    """1688 상품 링크 정규화 — offer ID만 있으면 그것을 키로."""
    if not url or not isinstance(url, str) or not url.startswith("http"):
        return ""
    m = _OFFER_RE.search(url)
    if m:
        return "offer:" + m.group(1)
    u = url.split("?")[0].split("#")[0].strip("/").lower()
    u = re.sub(r"^https?://", "", u)
    return "url:" + u


def _norm_img(url):
    """alicdn 이미지 URL 정규화 — ibank ID (O1CN...)만 있으면 그것을 키로.
    사이즈/webp suffix 무시.
    """
    if not url or not isinstance(url, str) or not url.startswith("http"):
        return ""
    m = _IBANK_RE.search(url)
    if m:
        return "ibank:" + m.group(1)
    u = url.split("?")[0].split("#")[0].lower()
    u = re.sub(r"^https?://", "", u)
    u = re.sub(r"_\d+x\d+\.jpg.*$", ".jpg", u)
    u = re.sub(r"\.jpg_\.webp$", ".jpg", u)
    return "img:" + u


def aggregate_products(orders):
    """모든 아이템을 dedupe. link 또는 img 하나라도 겹치면 같은 상품으로 병합 (union-find).

    - link 정규화: 1688 offer ID 추출
    - img 정규화: alicdn ibank ID 추출 (사이즈/webp suffix 무시)
    - 둘 다 없으면 (kor|cn) 이름 조합 폴백
    - 병합 후 상품별 total_qty/order_count/price min/max/avg 집계
    """
    items_list = []  # (item, order_group, order_file)
    parent = []

    def find(x):
        while parent[x] != x:
            parent[x] = parent[parent[x]]
            x = parent[x]
        return x

    def union(x, y):
        rx, ry = find(x), find(y)
        if rx != ry:
            parent[rx] = ry

    link_seen = {}
    img_seen = {}
    name_seen = {}

    for o in orders:
        for it in o.get("items", []):
            idx = len(items_list)
            items_list.append((it, o.get("group") or "", o.get("file") or ""))
            parent.append(idx)

            lk = _norm_link(it.get("link"))
            ik = _norm_img(it.get("img"))
            if lk:
                if lk in link_seen:
                    union(idx, link_seen[lk])
                else:
                    link_seen[lk] = idx
            if ik:
                if ik in img_seen:
                    union(idx, img_seen[ik])
                else:
                    img_seen[ik] = idx
            if not lk and not ik:
                nm = ((it.get("kor") or "").strip() + "|" + (it.get("cn") or "").strip()).strip("|")
                if nm:
                    if nm in name_seen:
                        union(idx, name_seen[nm])
                    else:
                        name_seen[nm] = idx
                else:
                    # link/img/이름 다 없는 완전 빈 아이템은 제외
                    parent[idx] = -1  # mark as dropped

    # 그룹핑
    groups = {}
    for idx, (it, group, file) in enumerate(items_list):
        if parent[idx] == -1:
            continue
        root = find(idx)
        p = groups.get(root)
        if p is None:
            p = {
                "img": "",
                "link": "",
                "kor": "",
                "cn": "",
                "_qtys": [],
                "_prices": [],
                "_memos": [],
                "_groups": set(),
                "_files": set(),
            }
            groups[root] = p
        img = it.get("img") or ""
        link = it.get("link") or ""
        kor = it.get("kor") or ""
        cn = it.get("cn") or ""
        # 첫 번째 http URL 유지 (다른 URL은 무시)
        if not p["img"] and img.startswith("http"):
            p["img"] = img
        if not p["link"] and link.startswith("http"):
            p["link"] = link
        # 이름은 더 서술적인 것 우선 (긴 것)
        if len(kor) > len(p["kor"]):
            p["kor"] = kor
        if len(cn) > len(p["cn"]):
            p["cn"] = cn
        if it.get("qty") is not None:
            p["_qtys"].append(float(it["qty"]))
        if it.get("price_cny") is not None:
            p["_prices"].append(float(it["price_cny"]))
        memo = it.get("memo") or ""
        if memo and memo not in p["_memos"] and len(p["_memos"]) < 5:
            p["_memos"].append(memo)
        if group:
            p["_groups"].add(group)
        if file:
            p["_files"].add(file)

    result = []
    for p in groups.values():
        prices = p["_prices"]
        qtys = p["_qtys"]
        result.append({
            "img": p["img"],
            "link": p["link"],
            "kor": p["kor"],
            "cn": p["cn"],
            "total_qty": int(round(sum(qtys))) if qtys else None,
            "order_count": len(p["_files"]),
            "price_min": round(min(prices), 2) if prices else None,
            "price_max": round(max(prices), 2) if prices else None,
            "price_avg": round(sum(prices) / len(prices), 2) if prices else None,
            "groups": sorted(p["_groups"]),
            "memos": p["_memos"][:3],
        })
    result.sort(key=lambda x: (
        x.get("total_qty") or 0,
        x.get("order_count") or 0,
    ), reverse=True)
    return result


def main():
    files = sorted(FOLDER.glob("*.xls*"))
    files = [f for f in files if not f.name.startswith("~$") and not f.name.endswith(".part")]
    print(f"parsing {len(files)} files from {FOLDER}", file=sys.stderr, flush=True)

    orders = []
    errors = []
    for i, f in enumerate(files, 1):
        try:
            row = parse_workbook(f)
            if "error" in row:
                errors.append({"file": f.name, "error": row["error"]})
                print(f"  [{i}/{len(files)}] ERR {f.name}: {row['error']}", file=sys.stderr, flush=True)
                continue
            orders.append(row)
            if i % 20 == 0 or i == len(files):
                print(f"  [{i}/{len(files)}] {f.name[:40]}", file=sys.stderr, flush=True)
        except Exception as e:
            errors.append({"file": f.name, "error": str(e)})
            print(f"  [{i}/{len(files)}] EXC {f.name}: {e}", file=sys.stderr, flush=True)

    # 중복 제거: 정규화된 파일명별로 그룹핑, 아이템 많은 것 우선 (동수면 mtime 최신)
    dup_groups = {}
    for o in orders:
        key = dedupe_key(o["file"])
        dup_groups.setdefault(key, []).append(o)
    deduped = []
    dropped = 0
    for key, group in dup_groups.items():
        if len(group) == 1:
            deduped.append(group[0])
            continue
        # 아이템 갯수 desc → mtime desc
        group.sort(key=lambda x: (x.get("item_count") or 0, x.get("mtime") or ""), reverse=True)
        best = group[0]
        best["duplicates"] = [g["file"] for g in group[1:]]
        deduped.append(best)
        dropped += len(group) - 1
    print(f"  deduped: {len(orders)} → {len(deduped)} ({dropped} duplicates removed)", file=sys.stderr, flush=True)
    orders = deduped

    # 정렬: 아이템 있는 것 우선 → 날짜 desc → 파일명 desc
    orders.sort(key=lambda x: (
        1 if (x.get("item_count") or 0) > 0 else 0,
        x.get("date") or "0000-00-00",
        x.get("file", ""),
    ), reverse=True)

    # 아이템들을 link/img 기준으로 dedupe → 상품 카탈로그 생성
    products = aggregate_products(orders)
    print(f"  products: {sum(o['item_count'] for o in orders):,} items → {len(products):,} unique products", file=sys.stderr, flush=True)

    payload = {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "source_folder": str(FOLDER),
        "count": len(orders),
        "product_count": len(products),
        "errors": errors,
        "orders": orders,
        "products": products,
    }
    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(json.dumps(payload, ensure_ascii=False, indent=2))
    print(f"wrote {OUT} ({OUT.stat().st_size:,} bytes, {len(orders)} orders, {len(errors)} errors)", file=sys.stderr, flush=True)


if __name__ == "__main__":
    main()
